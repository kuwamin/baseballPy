import pandas as pd
from openpyxl import load_workbook
from modules.models import Batter, Pitcher, Player


def Aquire_data(file_path: str, team_name: str) -> tuple[list[Pitcher], list[Batter]]:
    """
    Excelファイルから指定されたチームの選手データを読み込み、インスタンス化する。

    Args:
        - file_path : 読み込み対象となるExcelファイルのパス
        - team_name : 取得対象のチーム名（シート名の接頭辞として使用）

    Returns:
        - tuple[List[Pitcher], List[Batter]]:
            - pitchers: 生成されたPitcherクラスのインスタンスリスト
            - batters: 生成されたBatterクラスのインスタンスリスト
    """
    wb = load_workbook(file_path, data_only=True)
    sheet_pitchers = wb[f"{team_name}_p"]
    sheet_batters = wb[f"{team_name}_b"]

    # ヘッダー行の読み込み
    headers_pitchers = [cell.value for cell in sheet_pitchers[1]]
    headers_batters = [cell.value for cell in sheet_batters[1]]

    pitchers = []
    batters = []

    # 投手データの取得
    for pitcher in sheet_pitchers.iter_rows(min_row=2, values_only=True):
        data_dict = dict(zip(headers_pitchers, pitcher))
        pitcher = Pitcher(data_dict)
        pitchers.append(pitcher)

    # 野手データの取得
    for batter in sheet_batters.iter_rows(min_row=2, values_only=True):
        data_dict = dict(zip(headers_batters, batter))
        batter = Batter(data_dict)
        batters.append(batter)

    return pitchers, batters


def output_exam(
    file_path: str, team_name_1: str, team_name_2: str, all_players: list[Player]
) -> None:
    """
    試合結果（統計データ）をExcelファイルに書き戻す（加算更新）

    Args:
        - file_path : 読み込み対象となるExcelファイルのパス
        - team_name_1 : 取得対象のチーム名（シート名の接頭辞として使用）
        - team_name_2 : 取得対象のチーム名（シート名の接頭辞として使用）
        - all_players : 全選手のインスタンス

    Returns:
        - None
    """
    excel_data = pd.read_excel(file_path, sheet_name=None)

    # Excel列名とプログラム内statsキーの紐付け
    mapping = {
        "試合数": "games",
        "打席": "pa",
        "打数": "ab",
        "安打": "hits",
        "単打": "singles",
        "二塁打": "doubles",
        "三塁打": "triples",
        "本塁打": "hr",
        "打点": "rbi",
        "四球": "walks",
        "死球": "hbp",
        "三振": "so",
        "得点圏打数": "risp_ab",
        "得点圏安打": "risp_hits",
        "登板数": "games",
        "先発数": "starts",
        "勝利": "wins",
        "敗北": "losses",
        "セーブ": "saves",
        "ホールド": "holds",
        "完投": "complete_games",
        "完封": "shutouts",
        "打者数": "bf",
        "被安打": "hits_allowed",
        "被本塁打": "hr_allowed",
        "与四球": "walks_allowed",
        "与死球": "hbp_allowed",
        "奪三振": "strikeouts",
        "失点": "失点",
        "自責点": "自責点",
        "QS": "qs",
        "HQS": "hqs",
        "得点圏被打数": "risp_bf",
        "得点圏被安打": "risp_hits_allowed",
    }

    for player in all_players:
        # 野手かどうかを判定
        is_batter = hasattr(player, "position")
        t_name = team_name_1 if player.team == team_name_1 else team_name_2
        target_sheet = f"{t_name}_{'b' if is_batter else 'p'}"

        if target_sheet in excel_data:
            df = excel_data[target_sheet]
            idx_list = df.index[df["名前"] == player.name]

            if not idx_list.empty:
                idx = idx_list[0]

                # 1. 通常項目の加算
                for col, key in mapping.items():
                    if col in df.columns and key in player.stats:
                        val = player.stats[key]
                        if val > 0:
                            current_val = (
                                0 if pd.isna(df.at[idx, col]) else df.at[idx, col]
                            )
                            df.at[idx, col] = current_val + val

                # 2. 蓄積疲労・減少体力の更新（上書き）
                df.at[idx, "蓄積疲労"] = player.accumulated_fatigue
                if not is_batter:
                    df.at[idx, "減少体力"] = player.fatigue_stamina

                    # イニング数の特殊計算 (n.1, n.2 表記)
                    if player.stats.get("outs_pitched", 0) > 0:
                        cur = (
                            0.0
                            if pd.isna(df.at[idx, "イニング数"])
                            else df.at[idx, "イニング数"]
                        )
                        total_outs = (
                            int(cur) * 3
                            + round((cur - int(cur)) * 10)
                            + player.stats["outs_pitched"]
                        )
                        df.at[idx, "イニング数"] = (total_outs // 3) + (
                            total_outs % 3 / 10.0
                        )

    # 保存
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, df in excel_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def reset_columns(df, cols):
    """
    存在する列のみ0リセット
    """
    cols_to_reset = [c for c in cols if c in df.columns]
    df[cols_to_reset] = 0


def reset_result(file_path: str) -> None:
    """
        Excelの全成績列を0にリセットする

    Args:
        - file_path : 読み込み対象となるExcelファイルのパス

    Returns:
        - None
    """
    excel_data = pd.read_excel(file_path, sheet_name=None)

    pitcher_cols = [
        "減少体力",
        "蓄積疲労",
        "登板数",
        "先発数",
        "勝利",
        "敗北",
        "セーブ",
        "ホールド",
        "イニング数",
        "完投",
        "完封",
        "打者数",
        "奪三振",
        "与四球",
        "与死球",
        "被本塁打",
        "被安打",
        "失点",
        "自責点",
        "QS",
        "HQS",
        "得点圏被打数",
        "得点圏被安打",
    ]
    batter_cols = [
        "蓄積疲労",
        "試合数",
        "打席",
        "打数",
        "安打",
        "単打",
        "二塁打",
        "三塁打",
        "本塁打",
        "打点",
        "四球",
        "死球",
        "三振",
        "犠打",
        "犠飛",
        "盗塁成功",
        "盗塁死",
        "併殺打",
        "得点圏打数",
        "得点圏安打",
    ]

    for sheet_name, data_frame in excel_data.items():
        if sheet_name.endswith("_p"):
            reset_columns(data_frame, pitcher_cols)
        elif sheet_name.endswith("_b"):
            reset_columns(data_frame, batter_cols)

    # Excelファイルのそれぞれのシートに一括で書き込む
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, data in excel_data.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)
