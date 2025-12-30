from openpyxl import load_workbook
from Player import Batter, Pitcher

# テストコード
def test(team_name):
    file_path = 'test.xlsx'
    wb = load_workbook(file_path, data_only=True)
    
    # シート名の定義
    sheet = [wb[team_name+'_p'], wb[team_name+'_b']]

    # 投手・野手のヘッダ情報を取得
    headers_p = [cell.value for cell in sheet[0][1]]
    headers_b = [cell.value for cell in sheet[1][1]]

    # インスタンスのリストを作成
    pitchers = []
    batters = []


    # 選手データの取得
    for row_values in sheet[0].iter_rows(min_row=2, values_only=True):

        # ヘッダーと行の値を組み合わせて辞書を作成
        data_dict = dict(zip(headers_p, row_values))

        p = Pitcher(data_dict)
        pitchers.append(p)

    for row_values in sheet[1].iter_rows(min_row=2, values_only=True):

        # ヘッダーと行の値を組み合わせて辞書を作成
        data_dict = dict(zip(headers_b, row_values))

        b = Batter(data_dict)
        batters.append(b)

    '''
    print("\n--- 投手リスト ---")
    print(pitchers)
    print(pitchers[0].name)

    print("\n--- 野手リスト ---")
    print(batters)
    print(batters[0].name)
    '''

    return pitchers, batters